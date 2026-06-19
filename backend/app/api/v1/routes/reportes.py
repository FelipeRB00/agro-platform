from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
from datetime import datetime
from app.db.database import get_db
from app.core.dependencies import require_rol
from app.models.usuario import Usuario
from app.models.lista_compra import ListaCompra
from app.models.cotizacion import Cotizacion
from app.models.proveedor import Proveedor
from app.models.agricultor import Agricultor

# ReportLab
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# OpenPyXL
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter(prefix="/reportes", tags=["Reportes"])

def get_historial(db: Session, agricultor_id: int):
    listas = db.query(ListaCompra).filter(
        ListaCompra.agricultor_id == agricultor_id,
        ListaCompra.estado == "cerrada"
    ).order_by(ListaCompra.creado_en.desc()).all()

    resultado = []
    for lista in listas:
        cotizacion = db.query(Cotizacion).filter(
            Cotizacion.lista_id == lista.id,
            Cotizacion.estado == "aceptada"
        ).first()
        if not cotizacion:
            continue
        proveedor = db.query(Proveedor).filter(
            Proveedor.id == cotizacion.proveedor_id
        ).first()
        total = sum(float(i.subtotal or 0) for i in cotizacion.items)
        resultado.append({
            "id": f"ORD-{lista.id:04d}",
            "fecha": lista.creado_en.strftime("%d/%m/%Y"),
            "titulo": lista.titulo,
            "proveedor": proveedor.nombre_empresa if proveedor else "Desconocido",
            "items": len(lista.items),
            "total": total
        })
    return resultado

# ─── PDF ────────────────────────────────────────────────────────────

@router.get("/historial/pdf")
def exportar_pdf(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_rol("agricultor"))
):
    agricultor = db.query(Agricultor).filter(
        Agricultor.usuario_id == current_user.id
    ).first()

    pedidos = get_historial(db, agricultor.id)
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40, leftMargin=40,
        topMargin=50, bottomMargin=40
    )

    styles = getSampleStyleSheet()
    elements = []

    # Título
    titulo_style = ParagraphStyle(
        'titulo', parent=styles['Title'],
        fontSize=18, textColor=colors.HexColor('#1a3a1a'),
        spaceAfter=4, alignment=TA_CENTER
    )
    sub_style = ParagraphStyle(
        'sub', parent=styles['Normal'],
        fontSize=10, textColor=colors.HexColor('#666666'),
        spaceAfter=20, alignment=TA_CENTER
    )

    elements.append(Paragraph("CultivaTech", titulo_style))
    elements.append(Paragraph(
        f"Historial de Pedidos — {current_user.nombre}",
        sub_style
    ))
    elements.append(Paragraph(
        f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        sub_style
    ))
    elements.append(Spacer(1, 12))

    if not pedidos:
        elements.append(Paragraph("No hay pedidos registrados.", styles['Normal']))
    else:
        # Tabla
        headers = ["N° Orden", "Fecha", "Lista", "Proveedor", "Ítems", "Total (CLP)"]
        data = [headers]
        total_general = 0

        for p in pedidos:
            data.append([
                p["id"],
                p["fecha"],
                p["titulo"][:30] + "..." if len(p["titulo"]) > 30 else p["titulo"],
                p["proveedor"],
                str(p["items"]),
                f"${p['total']:,.0f}"
            ])
            total_general += p["total"]

        # Fila de total
        data.append(["", "", "", "", "TOTAL", f"${total_general:,.0f}"])

        col_widths = [80, 70, 150, 110, 45, 90]
        tabla = Table(data, colWidths=col_widths, repeatRows=1)
        tabla.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a1a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            # Filas
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f4f8f2')]),
            ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
            ('TOPPADDING', (0, 1), (-1, -2), 7),
            ('BOTTOMPADDING', (0, 1), (-1, -2), 7),
            # Fila total
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f4e8')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('TOPPADDING', (0, -1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#1a3a1a')),
        ]))
        elements.append(tabla)
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(
            f"Total de pedidos: {len(pedidos)}  |  Gasto total: ${total_general:,.0f} CLP",
            ParagraphStyle('resumen', parent=styles['Normal'],
                fontSize=10, textColor=colors.HexColor('#1a3a1a'),
                alignment=TA_LEFT)
        ))

    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=historial_pedidos_{datetime.now().strftime('%Y%m%d')}.pdf"
        }
    )

# ─── EXCEL ──────────────────────────────────────────────────────────

@router.get("/historial/excel")
def exportar_excel(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_rol("agricultor"))
):
    agricultor = db.query(Agricultor).filter(
        Agricultor.usuario_id == current_user.id
    ).first()

    pedidos = get_historial(db, agricultor.id)
    buffer = BytesIO()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial de Pedidos"

    # Colores
    verde_oscuro = "1a3a1a"
    verde_claro = "f4f8f2"
    verde_medio = "e8f4e8"

    # Título
    ws.merge_cells("A1:F1")
    ws["A1"] = "CultivaTech — Historial de Pedidos"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=verde_oscuro)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Agricultor: {current_user.nombre}  |  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Headers
    headers = ["N° Orden", "Fecha", "Lista de Compra", "Proveedor", "Ítems", "Total (CLP)"]
    header_row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=verde_oscuro)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            bottom=Side(style="medium", color=verde_oscuro)
        )
    ws.row_dimensions[header_row].height = 28

    # Datos
    total_general = 0
    for idx, p in enumerate(pedidos):
        row = header_row + 1 + idx
        fill_color = "FFFFFF" if idx % 2 == 0 else verde_claro

        valores = [p["id"], p["fecha"], p["titulo"], p["proveedor"], p["items"], p["total"]]
        for col, valor in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal="center" if col != 3 else "left", vertical="center")
            cell.font = Font(size=10)
            if col == 6:
                cell.number_format = '#,##0'
        ws.row_dimensions[row].height = 22
        total_general += p["total"]

    # Fila total
    total_row = header_row + 1 + len(pedidos)
    ws.merge_cells(f"A{total_row}:E{total_row}")
    ws[f"A{total_row}"] = f"TOTAL ({len(pedidos)} pedidos)"
    ws[f"A{total_row}"].font = Font(bold=True, size=11, color=verde_oscuro)
    ws[f"A{total_row}"].fill = PatternFill("solid", fgColor=verde_medio)
    ws[f"A{total_row}"].alignment = Alignment(horizontal="right", vertical="center")

    ws[f"F{total_row}"] = total_general
    ws[f"F{total_row}"].font = Font(bold=True, size=11, color=verde_oscuro)
    ws[f"F{total_row}"].fill = PatternFill("solid", fgColor=verde_medio)
    ws[f"F{total_row}"].number_format = '#,##0'
    ws[f"F{total_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[total_row].height = 26

    # Anchos de columna
    anchos = [12, 12, 35, 25, 8, 16]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=historial_pedidos_{datetime.now().strftime('%Y%m%d')}.xlsx"
        }
    )
    
# ═══════════════════════════════════════════════════════════════════
# ENDPOINT ORDEN DE COMPRA (reemplaza al antiguo comprobante_pdf)
# Reemplaza la función comprobante_pdf en reportes.py por esta.
# La ruta sigue siendo /comprobante/{cotizacion_id} para no romper el frontend,
# pero ahora genera una ORDEN DE COMPRA.
# ═══════════════════════════════════════════════════════════════════

@router.get("/comprobante/{cotizacion_id}")
def orden_compra_pdf(
    cotizacion_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_rol("agricultor"))
):
    from app.models.cotizacion import Cotizacion
    from app.models.lista_compra import ListaCompra, ItemLista
    from app.models.proveedor import Proveedor
    from app.models.insumo import Insumo
    from app.models.comision import Comision
    from datetime import timedelta

    agricultor = db.query(Agricultor).filter(
        Agricultor.usuario_id == current_user.id
    ).first()

    cotizacion = db.query(Cotizacion).filter(Cotizacion.id == cotizacion_id).first()
    if not cotizacion:
        raise HTTPException(status_code=404, detail="Cotización no encontrada")

    lista = db.query(ListaCompra).filter(ListaCompra.id == cotizacion.lista_id).first()
    proveedor = db.query(Proveedor).filter(Proveedor.id == cotizacion.proveedor_id).first()

    # Buscar la comisión asociada para saber el método de pago elegido
    comision = db.query(Comision).filter(
        Comision.cotizacion_id == cotizacion.id
    ).first()
    metodo_pago = comision.metodo_pago if comision else "contado"
    fecha_vencimiento = comision.fecha_vencimiento if comision else None

    items = []
    for i in cotizacion.items:
        item_lista = db.query(ItemLista).filter(ItemLista.id == i.item_lista_id).first()
        insumo = db.query(Insumo).filter(Insumo.id == item_lista.insumo_id).first() if item_lista else None
        items.append({
            "nombre": insumo.nombre if insumo else "",
            "ingrediente_activo": insumo.ingrediente_activo if insumo else None,
            "precio": float(i.precio_unitario),
            "cantidad": float(i.cantidad_ofrecida),
            "subtotal": float(i.subtotal or 0)
        })

    # Total sin IVA ni comisión al agricultor (precio limpio)
    subtotal = sum(it["subtotal"] for it in items)
    total = subtotal

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=50, bottomMargin=40)

    styles = getSampleStyleSheet()
    elements = []

    titulo_style = ParagraphStyle('t', parent=styles['Title'], fontSize=20,
        textColor=colors.HexColor('#1a3a1a'), spaceAfter=4, alignment=TA_CENTER)
    sub_style = ParagraphStyle('s', parent=styles['Normal'], fontSize=11,
        textColor=colors.HexColor('#666666'), spaceAfter=4, alignment=TA_CENTER)

    elements.append(Paragraph("CultivaTech", titulo_style))
    elements.append(Paragraph("ORDEN DE COMPRA", ParagraphStyle('oc', parent=styles['Title'],
        fontSize=14, textColor=colors.HexColor('#2d6a2d'), spaceAfter=4, alignment=TA_CENTER)))
    elements.append(Spacer(1, 20))

    # Datos de la orden
    metodo_texto = "Contado" if metodo_pago == "contado" else f"Crédito ({cotizacion.dias_credito or 0} días)"
    info_data = [
        ["Orden de Compra N°:", f"OC-{cotizacion.id:05d}"],
        ["Fecha de emisión:", datetime.now().strftime('%d/%m/%Y %H:%M')],
        ["Comprador (Agricultor):", current_user.nombre],
        ["Vendedor (Proveedor):", proveedor.nombre_empresa if proveedor else "—"],
        ["Lista de compra:", lista.titulo if lista else "—"],
        ["Método de pago:", metodo_texto],
    ]

    # Si es crédito, agregar fecha de vencimiento
    if metodo_pago == "credito" and fecha_vencimiento:
        info_data.append(["Fecha compromiso de pago:", fecha_vencimiento.strftime('%d/%m/%Y')])

    info_table = Table(info_data, colWidths=[160, 310])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1a3a1a')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 10))

    # Destacar el método de pago en un recuadro
    metodo_color = '#e8f4e8' if metodo_pago == "contado" else '#e3f0fb'
    metodo_borde = '#2d6a2d' if metodo_pago == "contado" else '#1a6db5'
    metodo_label = "PAGO AL CONTADO" if metodo_pago == "contado" else f"PAGO A CRÉDITO — {cotizacion.dias_credito or 0} DÍAS"
    metodo_box = Table([[metodo_label]], colWidths=[470])
    metodo_box.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(metodo_color)),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor(metodo_borde)),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor(metodo_borde)),
    ]))
    elements.append(metodo_box)
    elements.append(Spacer(1, 20))

    # Tabla items (con ingrediente activo)
    data = [["Producto", "Ingrediente Activo", "Precio Unit.", "Cant.", "Subtotal"]]
    for it in items:
        data.append([
            it["nombre"][:25],
            (it["ingrediente_activo"] or "—")[:22],
            f"${it['precio']:,.0f}",
            f"{it['cantidad']:,.0f}",
            f"${it['subtotal']:,.0f}"
        ])

    tabla = Table(data, colWidths=[130, 120, 80, 50, 90])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a1a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (1, -1), 'LEFT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f4f8f2')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
    ]))
    elements.append(tabla)
    elements.append(Spacer(1, 20))

    # Total (sin IVA ni comisión)
    totales_data = [
        ["TOTAL ORDEN:", f"${total:,.0f} CLP"],
    ]
    totales_table = Table(totales_data, colWidths=[330, 140])
    totales_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 14),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1a3a1a')),
        ('LINEABOVE', (0, 0), (-1, -1), 1.5, colors.HexColor('#1a3a1a')),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(totales_table)
    elements.append(Spacer(1, 40))

    # Pie de página
    pie = ParagraphStyle('pie', parent=styles['Normal'], fontSize=8,
        textColor=colors.HexColor('#999999'), alignment=TA_CENTER)
    elements.append(Paragraph(
        "Esta Orden de Compra es un documento generado por CultivaTech para formalizar el acuerdo "
        "entre comprador y vendedor. El pago se gestiona directamente entre las partes a través de los "
        "sistemas administrativos de la empresa. No constituye una boleta o factura electrónica ante el SII.",
        pie
    ))

    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=orden_compra_OC-{cotizacion.id:05d}.pdf"
        }
    )